# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://doc.scrapy.org/en/latest/topics/item-pipeline.html
import os
from openpyxl import Workbook
from openpyxl import load_workbook

class UrlscrapyPipeline(object):
    if os.path.exists('./tuniu.xlsx'):
        wb = load_workbook("./tuniu.xlsx")
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(['网站名称', '首页网址', '主办单位', '单位性质', '网站备案/许可证号', '服务器IP', 'IP所在地', '协议类型', '服务器类型', '页面类型', '域名', '注册商',
                   'WHOIS服务器', 'DNF服务器', '创建时间', '全球排名（PV Rank）', '访客排名（UV Rank）', '国家/地区', '国家/地区排名', '排名走势图链接',
                   '搜索流量占比图链接'])  # 设置表头

    def process_item(self, item, spider):
        line = [item['web_name'], item['web_home'], item['com_name'], item['icp_type'], item['icp_no'],
                item['server_ip'], item['server_location'], item['http_type'], item['server_type'], item['content_type'],
                item['domain'], item['reg_server'], item['server'], item['nserver'], item['create_day'],
                item['world_rank'], item['world_uv_rank'], item['country_code'], item['country_rank'],
                item['rank_trend_chart'], item['search_proportion_chart']]  # 把数据中每一项整理出来
        self.ws.append(line)  # 将数据以行的形式添加到xlsx中
        self.wb.save('./tuniu.xlsx')  # 保存xlsx文件
        print("OKKK---------------------------------------------------------------------")
        return item


